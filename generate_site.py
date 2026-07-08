import json,re,shutil,datetime as dt
from pathlib import Path
from openpyxl import load_workbook
EXCEL_FILE='WK2026_poule_beheer.xlsx'; SRC_DIR=Path('src'); SITE_DIR=Path('site')
months={'januari':1,'februari':2,'maart':3,'april':4,'mei':5,'juni':6,'juli':7,'augustus':8,'september':9,'oktober':10,'november':11,'december':12}
def time_to_str(v):
    if v is None:return ''
    if isinstance(v,dt.datetime):return v.strftime('%H:%M')
    if isinstance(v,dt.time):return v.strftime('%H:%M')
    return str(v)
def parse_dutch_date(text):
    if not isinstance(text,str):return None
    m=re.search(r'(\d{1,2})\s+([a-zé]+)\s+(\d{4})',text.lower())
    return dt.date(int(m.group(3)),months[m.group(2)],int(m.group(1))) if m else None
def tdf(wb):
    if 'Stand TdF' not in wb.sheetnames:return []
    ws=wb['Stand TdF']; out=[]
    for r in range(3,ws.max_row+1):
        name=ws.cell(r,2).value
        if not name:continue
        rank=ws.cell(r,1).value; total=ws.cell(r,3).value
        out.append({'rang':int(rank) if isinstance(rank,(int,float)) else rank,'deelnemer':str(name),'totaal':int(total) if isinstance(total,(int,float)) else (total or 0)})
    return out
def extract_data(excel_file):
    wb=load_workbook(excel_file,data_only=True)
    ws=wb['Stand']; stand=[]; participants=[]
    for r in range(3,ws.max_row+1):
        name=ws.cell(r,2).value
        if not name:continue
        rank=ws.cell(r,1).value
        stand.append({'rang':int(rank) if isinstance(rank,(int,float)) and rank is not None else rank,'deelnemer':str(name),'punten_wedstrijden':ws.cell(r,3).value or 0,'punten_bonus':ws.cell(r,4).value or 0,'totaal':ws.cell(r,5).value or 0}); participants.append(str(name))
    ws=wb['Bonus_beheer']; bp=[]
    for c in range(4,ws.max_column+1,2):
        name=ws.cell(1,c).value
        if name: bp.append((c,str(name)))
    bonus=[]
    for r in range(3,ws.max_row+1):
        nr=ws.cell(r,1).value; vraag=ws.cell(r,2).value; corr=ws.cell(r,3).value
        if vraag=='Totaal (bonus)':continue
        if vraag:
            answers={}; points={}
            for c,name in bp: answers[name]=ws.cell(r,c).value; points[name]=ws.cell(r,c+1).value
            bonus.append({'nr':nr,'vraag':str(vraag),'correct_antwoord':'' if corr is None else str(corr),'antwoorden':answers,'punten':points})
    ws=wb['Wedstrijden_beheer']; wp=[]
    for c in range(5,ws.max_column+1,2):
        name=ws.cell(1,c).value
        if name and not str(name).startswith('_'): wp.append((c,str(name)))
    phase={'Groepsfase','Zestiende finales','Achtste finales','Kwartfinales','Halve finales','Troostfinale','Finale','(Verliezers)finale','Verliezersfinale'}; stage=''; date_label=''; date_iso=''; matches=[]
    for r in range(3,ws.max_row+1):
        a,b,c,d=ws.cell(r,1).value,ws.cell(r,2).value,ws.cell(r,3).value,ws.cell(r,4).value
        if isinstance(a,str) and b is None and c is None and d is None:
            if a.strip() in phase or a.strip().lower() in {p.lower() for p in phase}: stage=a.strip()
            else:
                date_label=a.strip(); parsed=parse_dutch_date(date_label); date_iso=parsed.isoformat() if parsed else ''
            continue
        if b is not None and d is not None:
            predictions={}; points={}
            for col,name in wp: predictions[name]=ws.cell(r,col).value; points[name]=ws.cell(r,col+1).value
            matches.append({'fase':stage,'datum':date_label,'datum_iso':date_iso,'tijd':time_to_str(a),'thuis':str(b),'uitslag':'' if c is None else str(c),'uit':str(d),'predictions':predictions,'points':points})
    return {'meta':{'titel':'BINX Poules','bronbestand':EXCEL_FILE,'laatst_ververst':dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),'layout':'wk-tdf-tabs-v6'},'participants':participants,'stand':stand,'tdf_stand':tdf(wb),'bonus_questions':bonus,'matches':matches}
def build_site():
    data=extract_data(EXCEL_FILE); SITE_DIR.mkdir(exist_ok=True)
    for fn in ['index.html','style.css','app.js']: shutil.copy2(SRC_DIR/fn,SITE_DIR/fn)
    (SITE_DIR/'data.js').write_text('window.POULE_DATA = '+json.dumps(data,ensure_ascii=False,indent=2)+';\n',encoding='utf-8')
if __name__=='__main__': build_site()
